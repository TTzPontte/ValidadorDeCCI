#Importar as bibliotecas
from xmlrpc.client import DateTime
import openpyxl as xl
import pandas as pd
import sys
import os
import PyPDF2
import re
from openpyxl import load_workbook
import xlsxwriter
from time import sleep
import win32com.client as win32
import xlwings as xw
from datetime import datetime
import warnings

warnings.simplefilter(action='ignore', category=FutureWarning)

file_name = os.path.basename(sys.path[0])
pathBruto = sys.path[0]

posicaoFinal = len(pathBruto) - len(file_name) - 1

pathTratado = (pathBruto[0:posicaoFinal]).capitalize()
pathTratado

#Lógica para localizar arquivo na pasta .PDF e .XLSX <--- Contrato

#Função Que Acha Arquivo
def caminhoArquivo(pathProcurar, palavra1, palavra2, palavra3, palavra4, extensao):
    for root, dirs, files in os.walk(pathProcurar):
        for file in files:
            if palavra1 in file or palavra2 in file or palavra3 in file or palavra4 in file:
                if extensao in file:
                    pathArquivo = os.path.join(root, file)
                    return pathArquivo

def atribCarencia(x):
    if x > 0:
        x = x - 1
    return x



#Palavras Chaves
palavraChave1 = 'Contrato'
palavraChave2 = 'contrato'
palavraChave3 = 'CONTRATO'
palavraChave4 = 'Assinatura'
palavraChave5 = 'Cálculo'
palavraChave6 = 'Calculo'
palavraChave7 = 'Kit'
palavraChave8 = 'calculo'

#Extensão do Arquvo
xlsx = '.xlsx'
pdf = '.pdf'

arquivoPDF = caminhoArquivo(pathTratado, palavraChave1, palavraChave2, palavraChave3, palavraChave4, pdf)
arquivoXLSX = caminhoArquivo(pathTratado, palavraChave5, palavraChave6, palavraChave7, palavraChave8, xlsx)


# print("Caminho da planilha:", arquivoXLSX)
# print("\nCaminho do PDF:", arquivoPDF)

### Abre o arquivo pdf e xlsx ###
path = arquivoPDF
pdf_file = open(path, 'rb')

#Faz a leitura usando a biblioteca
read_pdf = PyPDF2.PdfFileReader(pdf_file)

# pega o numero de páginas
number_of_pages = read_pdf.getNumPages()

#Extriar Texto Página 1 a 5
text=''
for i in range(0,5):
    #Ler Página PDF
    pageObj = read_pdf.getPage(i)
    #Extrair Texto
    text=text+pageObj.extractText()

#Tratar Texto (Remover Quebra de Linhas)
text = re.sub('\r', '', text) 
text = re.sub('\n', '', text)
text = re.sub('\s+', ' ', text)
text = re.sub(' {2,}', ' ', text).strip(' ')
text = re.sub(' :', ':', text)

path2 = arquivoXLSX
base = xl.load_workbook(path2, data_only=True)

if 'financiamento' in path:
    produto = 'FI'
    
    try:
        planilha = base["Sim_PRICE"] # Abrir Aba
    except:
        planilha = base["Sim_SAC"] # Abrir Aba


    #Pegar valores da planilha
    valorTotal = round(planilha.cell(row=2,column=3).value, 2)
    tabela = planilha.cell(row=3,column=3).value.upper()
    try:
        iofCorrigido = round(planilha.cell(row=4,column=3).value, 2)
    except:
        iofCorrigido = 0
    tag = round(planilha.cell(row=5,column=3).value, 2)
    registro = round(planilha.cell(row=6,column=3).value, 2)
    valorLiquido = round(planilha.cell(row=7,column=3).value, 2)
    prazoMes = planilha.cell(row=8,column=3).value
    taxa = planilha.cell(row=9,column=3).value
    taxa = round((((taxa + 1) ** (12/1))-1), 4) # Passando a taxa de a.a para a.m
    primeiraParcela = round(planilha.cell(row=10,column=3).value, 2)
    valorImóvel = round(planilha.cell(row=11,column=3).value, 2)
    carencia = planilha.cell(row=12,column=3).value
    mesNpaga = planilha.cell(row=13,column=3).value
    cet = round(planilha.cell(row=14,column=3).value, 4)
    prazoContrato = prazoMes - carencia
    try:
        iofTotal = round(planilha.cell(row=15,column=18).value, 2)
    except:
        iofTotal = 0
    diaPagto = planilha.cell(row=3,column=6).value
    dataContrato = planilha.cell(row=2,column=6).value
    dataContrato = dataContrato.strftime('%d/%m/%Y')
    dataContratoTab = planilha.cell(row=18,column=2).value
    dataContratoTab = dataContratoTab.strftime('%d/%m/%Y')
    # Validando a data do contrato
    if dataContrato != dataContratoTab:
        dataContrato = 'ERRO Diferente da Tabela (B:18)'

    # Validando a data da ultima parcela 
    cont = 25
    ultimaParcela = planilha.cell(row=cont,column=2).value
    while ultimaParcela != None:
        cont = cont + 1
        ultimaParcela = planilha.cell(row=cont,column=2).value
    ultimaParcela = planilha.cell(row=cont-1,column=2).value
    #ultimaParcela = datetime(ultimaParcela)
    ultimaParcela = ultimaParcela.strftime('%d/%m/%Y')
    # Validando Amortização com o campo Tabela
    amortizacaoRef1 = planilha.cell(row=cont-1, column=3).value
    amortizacaoRef = planilha.cell(row=cont-2, column=3).value
    if amortizacaoRef1 == amortizacaoRef:
        amortizacao = 'SAC'
    else:
        amortizacao = 'Price'

        #Criar DF a partir do Dicionario XLSX
    listaDePara = {'valorTotal':valorTotal,'tabela': tabela,'iofCorrigido': iofCorrigido,
                    'tag': tag,'registro': registro,
                    'valorLiquido': valorLiquido,'prazoMes': prazoMes,
                    'taxa': taxa, 'primeiraParcela':primeiraParcela,
                    'valorImóvel':valorImóvel,'cet':cet,'prazoContrato': prazoContrato,
                    'ultimaParcela': ultimaParcela,
                    'dataContrato': dataContrato, 'carencia': carencia, 'diaPagto': diaPagto
                    }

    listaKey = []
    listaValues = []


    for key, value in listaDePara.items():
        listaKey.append(key)
        listaValues.append(value)


    #Criar Dicionario das duas Listas
    dict_keyValue = dict(zip(listaKey,listaValues))



    #Criar DF a partir do Dicionario PFD
    df_1 = pd.DataFrame([dict_keyValue])

    listaDePara = {'valorTotal':'Valor do Financiamento: R$','tabela': 'Sistema de Amortização:','iofCorrigido':'IOF: R$',
                    'tag': 'Abertura de Cadastro: R$','registro': 'Despesas de Registro (estimado): R$',
                    'valorLiquido': 'Valor Líquido a Liberar do Financiamento: R$','prazoMes': 'PRAZO DE AMORTIZAÇÃO:',
                    'taxa': 'exponencial ao mês, equivalente a de', 'primeiraParcela':'VALOR TOTAL DO PRIMEIRO ENCARGO, NESTA DATA: R$',
                    'valorImóvel':'leilão: R$','cet':'Custo Efetivo Total (CET):','prazoContrato': 'N.º DE PRESTAÇÕES:',
                    'ultimaParcela':'DATA DE VENCIMENTO DA ÚLTIMA PRESTAÇÃO:',
                    'dataContrato': 'Data de Desembolso:',
                    }

    listaKey = []
    listaValues = []

    for key, value in listaDePara.items():
        inicioFrase = text.find(value,0)
        finalFrase = inicioFrase + len(value) + 1
        proximoEspaco = text.find(" ", finalFrase)
        valorExtraido = text[finalFrase:proximoEspaco]

        if '/' in valorExtraido:
            valorExtraido = valorExtraido.replace(",", "")

        #Ajustar Valores Númericos
        if '.' in valorExtraido:
            valorExtraido = valorExtraido.replace(".", "")
            valorExtraido = valorExtraido.replace(",", ".")

        if ',' in valorExtraido:
            valorExtraido = valorExtraido.replace(".", "")
            valorExtraido = valorExtraido.replace(",", ".")
        
        #Ajustar Valores Percentuais
        if '%' in valorExtraido:
            valorExtraido = valorExtraido.replace(",", ".")
            valorExtraido = valorExtraido.replace("%", "")
            valorExtraido = round(float(valorExtraido)/100,4)
        
        
        listaKey.append(key)
        listaValues.append(valorExtraido)
        #Criar Dicionario das duas Listas
    dict_keyValue = dict(zip(listaKey,listaValues))

    #Criar DF a partir do Dicionario
    df_2 = pd.DataFrame([dict_keyValue])
    #Definir Variáveis Auxiliares
    topico4 = '4. Despesas'
    topico5 = '5. Valor Destinado'

    #Pegar posição das variáveis auxiliares no texto
    inicioTopico = text.find(topico4, 0)
    finalTopico = text.find(topico5, 0)

    #Criar Paragráfo Auxiliar (Somente com os sub itens do tópico 4. Despesas)
    paragrafo4 = text[inicioTopico+len(topico4)+1:finalTopico-1]
    paragrafo4 = re.sub('\s+',' ', paragrafo4)
    paragrafo4

    listaChave = ['4.1.', '4.2.','4.3.','4.4.', '4.5.']
    inicioItens = []

    for item in listaChave:
        inicioP = paragrafo4.find(item,0)
        inicioItens.append(inicioP)
        
    item1 = paragrafo4[inicioItens[0]:inicioItens[1]-1]
    item2 = paragrafo4[inicioItens[1]:inicioItens[2]-1]
    item3 = paragrafo4[inicioItens[2]:inicioItens[3]-1]
    try:
        item4 = paragrafo4[inicioItens[3]:inicioItens[4]-1]
    except:
        item4 = paragrafo4[inicioItens[3]:len(paragrafo4)]
    try:
        item5 = paragrafo4[inicioItens[4]:len(paragrafo4)]
    except:
        item5 = "N/A"

    listaFinal = [item1, item2, item3, item4, item5]

    listaValor = []

    for itemAux in listaFinal:
        if '[X]' in itemAux:
            inicioAux = itemAux.find('R$ ', 0)
            fimAux = itemAux.find(",", inicioAux) + 3
            resultadoAux = itemAux[inicioAux+3:fimAux]
        else:
            resultadoAux = '0,00'

        listaValor.append(resultadoAux)

    #Criar Dicionario das duas Listas
    dict_chaveValor = dict(zip(listaChave,listaValor))

    df_2['registro'] = pd.to_numeric(df_2['registro'], errors= 'coerce')
    df_2['registro'] = 0
    for key, value in dict_chaveValor.items():
        print(key)
        if key == '4.2.':
            value = float(value.replace(".","").replace(",","."))
            df_2['registro'] = df_2['registro'] + value 
        if key == '4.3.':
            value = float(value.replace(".","").replace(",","."))
            df_2['registro'] = df_2['registro'] + value
        if key == '4.4.':
            value = float(value.replace(".","").replace(",","."))
            df_2['registro'] = df_2['registro'] + value
        if key == '4.5.':
            value = float(value.replace(".","").replace(",","."))
            df_2['registro'] = df_2['registro'] + value

elif 'home equity' in path:
    produto = 'HE'

    try:
        planilha = base["PRICE_CORR"] # Abrir Aba
    except:
        planilha = base["SAC_CORR"] # Abrir Aba

    print('lendo planilha')

    #Pegar valores da planilha
    valorTotal = round(planilha.cell(row=2,column=3).value, 2)
    tabela = planilha.cell(row=3,column=3).value.upper()
    iofCorrigido = round(planilha.cell(row=4,column=3).value, 2)
    tag = round(planilha.cell(row=5,column=3).value, 2)
    registro = round(planilha.cell(row=6,column=3).value, 2)
    valorLiquido = round(planilha.cell(row=7,column=3).value, 2)
    prazoMes = planilha.cell(row=8,column=3).value
    taxa = planilha.cell(row=9,column=3).value
    taxa = round((((taxa + 1) ** (12/1))-1), 4)
    primeiraParcela = round(planilha.cell(row=10,column=3).value, 2)
    valorImóvel = round(planilha.cell(row=11,column=3).value, 2)
    carencia = planilha.cell(row=12,column=3).value
    mesNpaga = planilha.cell(row=13,column=3).value
    cet = round(planilha.cell(row=14,column=3).value, 4)
    prazoContrato = planilha.cell(row=15,column=3).value
    prazoContrato = prazoContrato - carencia
    iofTotal = round(planilha.cell(row=15,column=18).value, 2)
    diaPagto = planilha.cell(row=3,column=6).value
    dataContrato = planilha.cell(row=2,column=6).value
    dataContrato = dataContrato.strftime('%d/%m/%Y')
    dataContratoTab = planilha.cell(row=18,column=2).value
    dataContratoTab = dataContratoTab.strftime('%d/%m/%Y')
    # Validando a data do contrato
    if dataContrato != dataContratoTab:
        dataContrato = 'ERRO Diferente da Tabela (B:18)'

    # Validando a data da ultima parcela 
    cont = 20
    ultimaParcela = planilha.cell(row=cont,column=2).value
    while ultimaParcela != '':
        cont = cont + 1
        ultimaParcela = planilha.cell(row=cont,column=2).value
    ultimaParcela = planilha.cell(row=cont-1,column=2).value
    ultimaParcela = ultimaParcela.strftime('%d/%m/%Y')
    # Validando Amortização com o campo Tabela
    amortizacaoRef1 = planilha.cell(row=cont-1, column=3).value
    amortizacaoRef = planilha.cell(row=cont-2, column=3).value
    if amortizacaoRef1 == amortizacaoRef:
        amortizacao = 'SAC'
    else:
        amortizacao = 'Price'

        #Criar DF a partir do Dicionario XLSX
    listaDePara = {'valorTotal':valorTotal,'tabela': tabela,'iofCorrigido': iofCorrigido,
                    'tag': tag,'registro': registro,
                    'valorLiquido': valorLiquido,'prazoMes': prazoMes,
                    'taxa': taxa, 'primeiraParcela':primeiraParcela,
                    'valorImóvel':valorImóvel,'cet':cet,'prazoContrato': prazoContrato,
                    'ultimaParcela': ultimaParcela,
                    'dataContrato': dataContrato, 'carencia': carencia, 'diaPagto': diaPagto
                    }

    listaKey = []
    listaValues = []


    for key, value in listaDePara.items():
        listaKey.append(key)
        listaValues.append(value)


    #Criar Dicionario das duas Listas
    dict_keyValue = dict(zip(listaKey,listaValues))

    #Criar DF a partir do Dicionario PFD
    df_1 = pd.DataFrame([dict_keyValue])

    listaDePara = {'valorTotal':'VALOR DO EMPRÉSTIMO: R$','tabela': 'SISTEMA DE AMORTIZAÇÃO:','iofCorrigido': 'IOF: R$',
                    'tag': 'TARIFA DE ABERTURA DE CADASTRO: R$','registro': 'DESPESAS DE REGISTRO: R$',
                    'valorLiquido': '-M-N): R$','prazoMes': 'PRAZO DE AMORTIZAÇÃO:',
                    'taxa': 'H.1. NOMINAL:', 'primeiraParcela':'T. VALOR TOTAL DO PRIMEIRO ENCARGO, NESTA DATA: R$',
                    'valorImóvel':'leilão: R$','cet':'CUSTO EFETIVO TOTAL (CET):','prazoContrato': 'N.º DE PRESTAÇÕES:',
                    'ultimaParcela':'DATA DO TÉRMINO DO PRAZO CONTRATUAL:',
                    'dataContrato': 'DATA DE DESEMBOLSO:','saldoDevedor':'SALDO DEVEDOR DO IMÓVEL: R$','valorDisponivel':'(O-P): R$'
                    }

    #len(listaHE) # <--- Qtd de Itens na Lista
    listaKey = []
    listaValues = []

    for key, value in listaDePara.items():
        inicioFrase = text.find(value,0)
        finalFrase = inicioFrase + len(value) + 1
        proximoEspaco = text.find(" ", finalFrase)
        valorExtraido = text[finalFrase:proximoEspaco]

        #Ajustar Valores Númericos
        if '.' in valorExtraido or ',' in valorExtraido:
            valorExtraido = valorExtraido.replace(".", "")
            valorExtraido = valorExtraido.replace(",", ".")
        
        #Ajustar Valores Percentuais
        if '%' in valorExtraido:
            valorExtraido = valorExtraido.replace(",", ".")
            valorExtraido = valorExtraido.replace("%", "")
            valorExtraido = round(float(valorExtraido)/100,4)
        
        
        listaKey.append(key)
        listaValues.append(valorExtraido)



    #Criar Dicionario das duas Listas
    dict_keyValue = dict(zip(listaKey,listaValues))

    #Criar DF a partir do Dicionario
    df_2 = pd.DataFrame([dict_keyValue])
    df_2['valorDisponivel'] = float(df_2['valorDisponivel'])
    df_2['saldoDevedor'] = float(df_2['saldoDevedor'])
    df_2['valorLiquido'] = float(df_2['valorLiquido'])
    valid = df_2['valorLiquido']-df_2['saldoDevedor']
    valid2 = df_2['valorDisponivel']
    
else:
    print(' **---- ERRO NO DIRETÓRIO (não foi possivel saber se é FI ou HE) -----**')

#Criar Colunas Dos Campos que estão faltando
df_2.insert(len(listaKey),"carencia", [''])
df_2.insert(len(listaKey)+1,"diaPagto", [''])
if produto == 'HE':
    if valid.all() == valid2.all():
            print("********FOOOI*********")
            del df_2['saldoDevedor']
            del df_2['valorDisponivel']
    else:
        for loopera in range(0,10):
            print('@-'*20)
            print('O SALDO DEVEDOR MENOS O VALOR LIQUIDO NÃO ESTA BATENDO COM O VALOR DIPONIVEL')
            print('@-'*20)

#Inserir Valores Nas Colunas
df_2['carencia'] = df_2['prazoMes'].astype(float) - df_2['prazoContrato'].astype(float) 


df_2['diaPagto'] = pd.DatetimeIndex(df_2['ultimaParcela']).strftime('%d/%m/%Y')
df_2['diaPagto'] = pd.DatetimeIndex(df_2['diaPagto']).day

# transformando em float 
df_2['valorTotal'] = pd.to_numeric(df_2['valorTotal'], errors= 'coerce')
df_2['iofCorrigido'] = pd.to_numeric(df_2['iofCorrigido'], errors= 'coerce')
df_2['tag'] = pd.to_numeric(df_2['tag'], errors= 'coerce')
df_2['registro'] = pd.to_numeric(df_2['registro'], errors= 'coerce')
df_1['registro'] = pd.to_numeric(df_2['registro'], errors= 'coerce')
df_2['valorLiquido'] = pd.to_numeric(df_2['valorLiquido'], errors= 'coerce')
df_2['prazoMes'] = pd.to_numeric(df_2['prazoMes'], errors= 'coerce')
df_2['taxa'] = pd.to_numeric(df_2['taxa'], errors= 'coerce')
df_2['primeiraParcela'] = pd.to_numeric(df_2['primeiraParcela'], errors= 'coerce')
df_2['valorImóvel'] = pd.to_numeric(df_2['valorImóvel'], errors= 'coerce')
df_2['cet'] = pd.to_numeric(df_2['cet'], errors= 'coerce')
df_2['prazoContrato'] = pd.to_numeric(df_2['prazoContrato'], errors= 'coerce')

# Armazena df em uma planilha diferente do mesmo arquivo
writer = pd.ExcelWriter(pathTratado+'\Resultado.xlsx', engine='xlsxwriter')
df_3 = df_1 == df_2
df_3 = df_3.append(df_1)
df_3 = df_3.append(df_2)
df_3.to_excel(writer,'Plan1')
writer.save()
writer.close()
#planilha.book = book

### Acionamento da Macro

#-- Abrir planilha
wb = xw.Book("G:\Drives compartilhados\Pontte\Operações\Automações\Projetos TESTES\Code de validação test\Central Formatação.xlsm").sheets[0]
sleep(3)
#-- Rodar Macro
wb.range('E10').value = pathTratado+"\Resultado.xlsx"
wb = xw.Book("G:\Drives compartilhados\Pontte\Operações\Automações\Projetos TESTES\Code de validação test\Central Formatação.xlsm")
macro = wb.macro("AddFormatacao")
macro()
sleep(5)
#-- Salvar e fechar
wb.save()
wb.close()